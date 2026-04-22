"""Inspect structure of both Excel files."""
import pandas as pd
from openpyxl import load_workbook

def inspect(path):
    print(f"\n{'='*70}")
    print(f"FILE: {path}")
    print('='*70)
    wb = load_workbook(path, data_only=True, read_only=True)
    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        try:
            df = pd.read_excel(path, sheet_name=sheet_name, nrows=5)
            print(f"Shape (first 5 rows): {df.shape}")
            print(f"Columns ({len(df.columns)}):")
            for i, col in enumerate(df.columns):
                print(f"  [{i}] {col!r}")
            print("First 3 rows:")
            print(df.head(3).to_string(max_colwidth=40))
        except Exception as e:
            print(f"Error reading: {e}")
    wb.close()

inspect("notej1_j2.xlsx")
inspect("PROPOSITION_REPARTITION_DR_35_EQUIPES__.xlsx")
